#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script để tạo bảng giá offer dịch vụ thiết kế website cho các quán ăn
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Tạo workbook mới
wb = openpyxl.Workbook()

# Sheet 1: Bảng giá các gói cơ bản
ws1 = wb.active
ws1.title = "Bang Gia Goi Co Ban"

# Định nghĩa các gói
packages = [
    {
        "Tên gói": "GÓI STARTER",
        "Giá": "3.000.000 VNĐ",
        "Thời gian": "7-10 ngày",
        "Tính năng": [
            "Thiết kế responsive (mobile, tablet, desktop)",
            "Tối đa 5 trang (Trang chủ, Giới thiệu, Menu, Liên hệ, Gallery)",
            "Tích hợp Google Maps",
            "Form liên hệ cơ bản",
            "Tích hợp Facebook, Instagram",
            "Tối ưu SEO cơ bản",
            "Bảo hành 3 tháng",
            "Hỗ trợ cập nhật nội dung 1 lần/tháng (3 tháng đầu)"
        ]
    },
    {
        "Tên gói": "GÓI PROFESSIONAL",
        "Giá": "5.500.000 VNĐ",
        "Thời gian": "10-15 ngày",
        "Tính năng": [
            "Tất cả tính năng gói Starter",
            "Tối đa 10 trang",
            "Hệ thống đặt bàn online",
            "Tích hợp thanh toán online (Momo, ZaloPay, VNPay)",
            "Quản lý menu động (thêm/sửa/xóa món)",
            "Gallery ảnh không giới hạn",
            "Tích hợp Google Reviews",
            "Chatbot Facebook Messenger",
            "Bảo hành 6 tháng",
            "Hỗ trợ cập nhật nội dung 2 lần/tháng (6 tháng đầu)",
            "Training sử dụng hệ thống"
        ]
    },
    {
        "Tên gói": "GÓI PREMIUM",
        "Giá": "10.000.000 VNĐ",
        "Thời gian": "15-20 ngày",
        "Tính năng": [
            "Tất cả tính năng gói Professional",
            "Không giới hạn số trang",
            "Thiết kế UI/UX chuyên nghiệp, độc quyền",
            "Hệ thống đặt bàn nâng cao (chọn bàn, thời gian)",
            "Tích hợp đặt món online (food ordering)",
            "Hệ thống quản lý khách hàng (CRM)",
            "Tích hợp Google Analytics nâng cao",
            "Email marketing tích hợp",
            "Multi-language (Tiếng Việt + Tiếng Anh)",
            "Bảo hành 12 tháng",
            "Hỗ trợ cập nhật nội dung không giới hạn (12 tháng đầu)",
            "Training và tư vấn marketing online"
        ]
    }
]

# Header cho bảng so sánh
headers = ["Tính năng", "Gói Starter", "Gói Professional", "Gói Premium"]
ws1.merge_cells('A1:D1')
ws1['A1'] = "BẢNG GIÁ DỊCH VỤ THIẾT KẾ WEBSITE CHO QUÁN ĂN"
ws1['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws1['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 40

# Ghi header
for col_num, header in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Điều chỉnh độ rộng cột
ws1.column_dimensions['A'].width = 50
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 30

# Danh sách tính năng để so sánh
features_comparison = [
    ("Thiết kế responsive", "✓", "✓", "✓"),
    ("Số trang tối đa", "5 trang", "10 trang", "Không giới hạn"),
    ("Tích hợp Google Maps", "✓", "✓", "✓"),
    ("Form liên hệ", "✓", "✓", "✓"),
    ("Tích hợp mạng xã hội", "✓", "✓", "✓"),
    ("Đặt bàn online", "✗", "✓", "✓"),
    ("Thanh toán online", "✗", "✓", "✓"),
    ("Quản lý menu động", "✗", "✓", "✓"),
    ("Đặt món online", "✗", "✗", "✓"),
    ("Hệ thống CRM", "✗", "✗", "✓"),
    ("Multi-language", "✗", "✗", "✓"),
    ("Thời gian bảo hành", "3 tháng", "6 tháng", "12 tháng"),
    ("Hỗ trợ cập nhật", "1 lần/tháng", "2 lần/tháng", "Không giới hạn"),
]

# Ghi dữ liệu so sánh
for row_num, (feature, starter, pro, premium) in enumerate(features_comparison, 3):
    for col_num, value in enumerate([feature, starter, pro, premium], 1):
        cell = ws1.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left" if col_num == 1 else "center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        if col_num == 1:
            cell.font = Font(bold=True)

# Thêm giá
row_num = len(features_comparison) + 4
ws1.merge_cells(f'A{row_num}:D{row_num}')
ws1[f'A{row_num}'] = "GIÁ"
ws1[f'A{row_num}'].font = Font(bold=True, size=14, color="FFFFFF")
ws1[f'A{row_num}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws1[f'A{row_num}'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[row_num].height = 30

row_num += 1
prices = ["", "3.000.000 VNĐ", "5.500.000 VNĐ", "10.000.000 VNĐ"]
for col_num, price in enumerate(prices, 1):
    cell = ws1.cell(row=row_num, column=col_num)
    cell.value = price
    cell.font = Font(bold=True, size=12, color="C00000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Sheet 2: Chi tiết từng gói
ws2 = wb.create_sheet("Chi Tiet Goi")
row = 1

for package in packages:
    # Tên gói
    ws2.merge_cells(f'A{row}:B{row}')
    ws2[f'A{row}'] = package["Tên gói"]
    ws2[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws2[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws2[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[row].height = 35
    row += 1
    
    # Giá và thời gian
    ws2[f'A{row}'] = "Giá:"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'] = package["Giá"]
    ws2[f'B{row}'].font = Font(bold=True, size=12, color="C00000")
    row += 1
    
    ws2[f'A{row}'] = "Thời gian hoàn thành:"
    ws2[f'A{row}'].font = Font(bold=True)
    ws2[f'B{row}'] = package["Thời gian"]
    row += 1
    
    # Tính năng
    ws2[f'A{row}'] = "Tính năng bao gồm:"
    ws2[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    for feature in package["Tính năng"]:
        ws2[f'A{row}'] = "•"
        ws2[f'B{row}'] = feature
        ws2[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 2  # Khoảng cách giữa các gói

ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 80

# Sheet 3: Tính năng bổ sung (Add-ons)
ws3 = wb.create_sheet("Tinh Nang Bo Sung")

ws3.merge_cells('A1:C1')
ws3['A1'] = "TÍNH NĂNG BỔ SUNG (ADD-ONS)"
ws3['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

# Header
headers_addon = ["STT", "Tính năng", "Giá"]
for col_num, header in enumerate(headers_addon, 1):
    cell = ws3.cell(row=2, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Danh sách tính năng add-on
addons = [
    ("Tích hợp đặt món online (food ordering)", "2.200.000 VNĐ"),
    ("Hệ thống đặt bàn nâng cao (chọn bàn, thời gian)", "1.300.000 VNĐ"),
    ("Tích hợp thanh toán online (Momo, ZaloPay, VNPay)", "1.800.000 VNĐ"),
    ("Chatbot Facebook Messenger tự động", "1.200.000 VNĐ"),
    ("Hệ thống quản lý khách hàng (CRM)", "2.500.000 VNĐ"),
    ("Tích hợp Google Reviews", "600.000 VNĐ"),
    ("Email marketing tích hợp", "1.800.000 VNĐ"),
    ("Multi-language (thêm ngôn ngữ)", "1.200.000 VNĐ"),
    ("Thiết kế logo chuyên nghiệp", "1.200.000 VNĐ"),
    ("Thiết kế menu PDF in ấn", "900.000 VNĐ"),
    ("Video giới thiệu quán (30 giây)", "2.200.000 VNĐ"),
    ("Tích hợp Google Analytics nâng cao", "900.000 VNĐ"),
    ("Tối ưu SEO nâng cao", "1.800.000 VNĐ"),
    ("Tích hợp Live Chat (Zalo, Facebook)", "1.000.000 VNĐ"),
    ("Hệ thống tích điểm, voucher online", "2.200.000 VNĐ"),
    ("App mobile (iOS/Android) cơ bản", "6.500.000 VNĐ"),
    ("Tích hợp POS system", "3.000.000 VNĐ"),
    ("Dịch vụ quản lý nội dung hàng tháng", "1.200.000 VNĐ/tháng"),
    ("Dịch vụ marketing online (SEO, Google Ads)", "2.500.000 VNĐ/tháng"),
    ("Hosting & Domain (1 năm)", "1.000.000 VNĐ/năm"),
]

# Ghi dữ liệu add-on
for row_num, (feature, price) in enumerate(addons, 3):
    # STT
    cell = ws3.cell(row=row_num, column=1)
    cell.value = row_num - 2
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Tính năng
    cell = ws3.cell(row=row_num, column=2)
    cell.value = feature
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Giá
    cell = ws3.cell(row=row_num, column=3)
    cell.value = price
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True, color="C00000")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Tô màu xen kẽ
    if row_num % 2 == 0:
        for col in [1, 2, 3]:
            ws3.cell(row=row_num, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 60
ws3.column_dimensions['C'].width = 25

# Sheet 4: Thông tin liên hệ và ưu đãi
ws4 = wb.create_sheet("Thong Tin Lien He")

row = 1
ws4.merge_cells(f'A{row}:B{row}')
ws4[f'A{row}'] = "THÔNG TIN LIÊN HỆ & ƯU ĐÃI ĐẶC BIỆT"
ws4[f'A{row}'].font = Font(bold=True, size=16, color="FFFFFF")
ws4[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws4[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[row].height = 40
row += 2

# Thông tin liên hệ
ws4[f'A{row}'] = "THÔNG TIN LIÊN HỆ:"
ws4[f'A{row}'].font = Font(bold=True, size=14, color="1F4E78")
row += 1

contact_info = [
    ("Email:", "info@example.com"),
    ("Hotline:", "0900 000 000"),
    ("Website:", "www.example.com"),
    ("Địa chỉ:", "TP. Hồ Chí Minh"),
]

for label, value in contact_info:
    ws4[f'A{row}'] = label
    ws4[f'A{row}'].font = Font(bold=True)
    ws4[f'B{row}'] = value
    row += 1

row += 2

# Ưu đãi
ws4[f'A{row}'] = "ƯU ĐÃI ĐẶC BIỆT:"
ws4[f'A{row}'].font = Font(bold=True, size=14, color="C00000")
row += 1

promotions = [
    "🎁 Giảm 10% khi đặt 2 gói trở lên",
    "🎁 Tặng miễn phí domain + hosting năm đầu (áp dụng gói Professional trở lên)",
    "🎁 Tặng thiết kế logo cơ bản (áp dụng gói Professional trở lên)",
    "🎁 Giảm 15% cho khách hàng đặt trong tháng này",
    "🎁 Tặng 1 tháng hỗ trợ cập nhật nội dung miễn phí",
    "🎁 Tư vấn marketing online miễn phí (1 buổi)",
]

for promo in promotions:
    ws4[f'A{row}'] = promo
    ws4[f'A{row}'].font = Font(size=11)
    ws4[f'A{row}'].alignment = Alignment(wrap_text=True)
    row += 1

row += 2

# Lưu ý
ws4[f'A{row}'] = "LƯU Ý:"
ws4[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
row += 1

notes = [
    "• Giá trên chưa bao gồm VAT (nếu có)",
    "• Thanh toán: 50% khi ký hợp đồng, 50% khi bàn giao",
    "• Thời gian bảo hành tính từ ngày bàn giao website",
    "• Hỗ trợ cập nhật nội dung trong thời gian bảo hành",
    "• Các tính năng add-on có thể được thêm vào bất kỳ gói nào",
    "• Giá có thể thay đổi tùy theo yêu cầu cụ thể của khách hàng",
]

for note in notes:
    ws4[f'A{row}'] = note
    ws4[f'A{row}'].alignment = Alignment(wrap_text=True)
    row += 1

ws4.column_dimensions['A'].width = 80
ws4.column_dimensions['B'].width = 40

# Lưu file
output_file = "/Users/anhmai/Desktop/F&B Doanh Nghiệp/data/Bang_Gia_Thiet_Ke_Website.xlsx"
wb.save(output_file)
print(f"Đã tạo file Excel thành công: {output_file}")
print(f"File bao gồm {len(wb.sheetnames)} sheets:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")

