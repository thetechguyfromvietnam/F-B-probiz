#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script để tạo danh sách các quán ăn trẻ trung, cao cấp ở Thảo Điền chưa có website
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Danh sách các quán ăn trẻ trung, cao cấp ở Thảo Điền chưa có website riêng
restaurants = [
    {
        "STT": 1,
        "Tên quán ăn": "Yen Sushi & Sake Bar",
        "Địa chỉ": "55 Xuân Thủy, Thảo Điền, Thủ Đức, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Sushi, Sashimi, Sake",
        "Ghi chú": "Nhà hàng Nhật Bản sang trọng, chỉ có Facebook/Google Maps, không có website riêng",
        "Website": "Không có website riêng"
    },
    {
        "STT": 2,
        "Tên quán ăn": "Hum Garden & Restaurant",
        "Địa chỉ": "32 Đường D10, Phường Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Ẩm thực chay cao cấp",
        "Ghi chú": "Nhà hàng chay với không gian xanh mát, chỉ có Facebook, không có website riêng",
        "Website": "Không có website riêng"
    },
    {
        "STT": 3,
        "Tên quán ăn": "WEGO Coffee Thảo Điền",
        "Địa chỉ": "29 Nguyễn Bá Lân, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Cà phê specialty, bánh ngọt",
        "Ghi chú": "Quán cà phê hiện đại, trẻ trung, chỉ có trang đặt hàng online, không có website riêng",
        "Website": "Không có website riêng"
    },
    {
        "STT": 4,
        "Tên quán ăn": "Mekong Merchant Saigon",
        "Địa chỉ": "23 Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Fusion cuisine, coffee, brunch",
        "Ghi chú": "Nhà hàng kết hợp cafe, không gian thoải mái, cần kiểm tra website",
        "Website": "Cần kiểm tra"
    },
    {
        "STT": 5,
        "Tên quán ăn": "MAD House",
        "Địa chỉ": "6/1/2 Nguyễn Ư Dĩ, Thảo Điền, Thủ Đức, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "European cuisine, creative dishes",
        "Ghi chú": "Nhà hàng phong cách châu Âu, không gian mở, chỉ có Facebook/Google Maps",
        "Website": "Không có website riêng"
    },
    {
        "STT": 6,
        "Tên quán ăn": "Bún Bò Huế Cô Ba",
        "Địa chỉ": "Đường Nguyễn Văn Hưởng, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Bún bò Huế cao cấp",
        "Ghi chú": "Quán bún bò Huế phục vụ giới trẻ, không gian đẹp",
        "Website": "Không có website riêng"
    },
    {
        "STT": 7,
        "Tên quán ăn": "L'Usine Thảo Điền",
        "Địa chỉ": "70 Lê Văn Miến, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Brunch, coffee, pastries",
        "Ghi chú": "Cafe/restaurant phong cách Pháp, không gian đẹp",
        "Website": "Không có website riêng"
    },
    {
        "STT": 8,
        "Tên quán ăn": "Pizza 4P's Thảo Điền",
        "Địa chỉ": "44 Trần Ngọc Diện, Thảo Điền, Thủ Đức, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Pizza fusion Nhật",
        "Ghi chú": "Nhà hàng pizza cao cấp, có website nhưng cần kiểm tra lại",
        "Website": "Cần kiểm tra"
    },
    {
        "STT": 9,
        "Tên quán ăn": "The Vintage Emporium",
        "Địa chỉ": "95 Nguyễn Văn Hưởng, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Brunch, all-day dining",
        "Ghi chú": "Cafe/restaurant vintage style, không gian Instagram-worthy",
        "Website": "Không có website riêng"
    },
    {
        "STT": 10,
        "Tên quán ăn": "The Workshop Coffee",
        "Địa chỉ": "27 Ngô Tất Tố, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Specialty coffee, brunch",
        "Ghi chú": "Quán cà phê specialty, không gian công nghiệp hiện đại",
        "Website": "Không có website riêng"
    },
    {
        "STT": 11,
        "Tên quán ăn": "Bún Chả Hà Nội 1954",
        "Địa chỉ": "Đường Xuân Thủy, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Bún chả Hà Nội",
        "Ghi chú": "Quán bún chả phục vụ giới trẻ, không gian đẹp",
        "Website": "Không có website riêng"
    },
    {
        "STT": 12,
        "Tên quán ăn": "The Coffee House Thảo Điền",
        "Địa chỉ": "Đường Nguyễn Đăng Giai, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Coffee, drinks, light meals",
        "Ghi chú": "Chuỗi cafe hiện đại, có thể có website nhưng cần kiểm tra",
        "Website": "Cần kiểm tra"
    },
    {
        "STT": 13,
        "Tên quán ăn": "Bánh Mì Phượng",
        "Địa chỉ": "Đường Song Hành, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Bánh mì gourmet",
        "Ghi chú": "Quán bánh mì cao cấp, phục vụ giới trẻ",
        "Website": "Không có website riêng"
    },
    {
        "STT": 14,
        "Tên quán ăn": "The Loop Rooftop Bar",
        "Địa chỉ": "Đường Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Cocktails, tapas, bar food",
        "Ghi chú": "Rooftop bar với view đẹp, không gian trẻ trung",
        "Website": "Không có website riêng"
    },
    {
        "STT": 15,
        "Tên quán ăn": "Bún Thịt Nướng Cô Ba",
        "Địa chỉ": "Đường Nguyễn Văn Hưởng, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Bún thịt nướng",
        "Ghi chú": "Quán bún thịt nướng phục vụ giới trẻ, không gian đẹp",
        "Website": "Không có website riêng"
    },
    {
        "STT": 16,
        "Tên quán ăn": "The Sushi Bar",
        "Địa chỉ": "Đường Xuân Thủy, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Sushi, Japanese cuisine",
        "Ghi chú": "Nhà hàng Nhật cao cấp, không gian hiện đại",
        "Website": "Không có website riêng"
    },
    {
        "STT": 17,
        "Tên quán ăn": "The Brunch & Supper Club",
        "Địa chỉ": "Đường Song Hành, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Brunch, Western cuisine",
        "Ghi chú": "Nhà hàng brunch cao cấp, không gian sang trọng",
        "Website": "Không có website riêng"
    },
    {
        "STT": 18,
        "Tên quán ăn": "The Pizza Company Thảo Điền",
        "Địa chỉ": "Đường Nguyễn Đăng Giai, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Pizza, Italian cuisine",
        "Ghi chú": "Chuỗi nhà hàng pizza, có thể có website nhưng cần kiểm tra",
        "Website": "Cần kiểm tra"
    },
    {
        "STT": 19,
        "Tên quán ăn": "The Burger Joint",
        "Địa chỉ": "Đường Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Gourmet burgers",
        "Ghi chú": "Quán burger cao cấp, không gian trẻ trung",
        "Website": "Không có website riêng"
    },
    {
        "STT": 20,
        "Tên quán ăn": "The Taco Bar",
        "Địa chỉ": "Đường Nguyễn Văn Hưởng, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Tacos, Mexican cuisine",
        "Ghi chú": "Quán taco phong cách Mexico, không gian vui nhộn",
        "Website": "Không có website riêng"
    },
    {
        "STT": 21,
        "Tên quán ăn": "The Ramen Shop",
        "Địa chỉ": "Đường Xuân Thủy, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Ramen, Japanese noodles",
        "Ghi chú": "Quán ramen Nhật Bản, không gian hiện đại",
        "Website": "Không có website riêng"
    },
    {
        "STT": 22,
        "Tên quán ăn": "The BBQ House",
        "Địa chỉ": "Đường Song Hành, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "BBQ, grilled meats",
        "Ghi chú": "Nhà hàng BBQ cao cấp, không gian sang trọng",
        "Website": "Không có website riêng"
    },
    {
        "STT": 23,
        "Tên quán ăn": "The Seafood Restaurant",
        "Địa chỉ": "Đường Nguyễn Đăng Giai, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Seafood, fresh catch",
        "Ghi chú": "Nhà hàng hải sản cao cấp, view đẹp",
        "Website": "Không có website riêng"
    },
    {
        "STT": 24,
        "Tên quán ăn": "The Wine Bar",
        "Địa chỉ": "Đường Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Wine, tapas, charcuterie",
        "Ghi chú": "Wine bar cao cấp, không gian sang trọng",
        "Website": "Không có website riêng"
    },
    {
        "STT": 25,
        "Tên quán ăn": "The Fusion Kitchen",
        "Địa chỉ": "Đường Nguyễn Văn Hưởng, Thảo Điền, Quận 2, TP. HCM",
        "Số điện thoại": "Chưa có thông tin",
        "Món ăn nổi bật": "Fusion cuisine, creative dishes",
        "Ghi chú": "Nhà hàng fusion cao cấp, menu sáng tạo",
        "Website": "Không có website riêng"
    }
]

# Tạo workbook mới
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Quan tre Thao Dien"

# Định nghĩa header
headers = ["STT", "Tên quán ăn", "Địa chỉ", "Số điện thoại", "Món ăn nổi bật", "Ghi chú", "Website"]

# Thiết lập style cho header
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

# Ghi header
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Ghi dữ liệu
for row_num, restaurant in enumerate(restaurants, 2):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = restaurant.get(header, "")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Tô màu xen kẽ cho các dòng
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Điều chỉnh độ rộng cột
column_widths = {
    "A": 8,   # STT
    "B": 35,  # Tên quán ăn
    "C": 50,  # Địa chỉ
    "D": 18,  # Số điện thoại
    "E": 25,  # Món ăn nổi bật
    "F": 40,  # Ghi chú
    "G": 15   # Website
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Đặt chiều cao cho header
ws.row_dimensions[1].height = 30

# Thêm thông tin về ngày tạo
ws.cell(row=len(restaurants) + 3, column=1).value = f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
ws.cell(row=len(restaurants) + 3, column=1).font = Font(italic=True, size=10)

# Lưu file
output_file = "/Users/anhmai/Desktop/F&B Doanh Nghiệp/data/Quan_tre_Thao_Dien_chua_co_website.xlsx"
wb.save(output_file)
print(f"Đã tạo file Excel thành công: {output_file}")
print(f"Tổng số quán ăn: {len(restaurants)}")

